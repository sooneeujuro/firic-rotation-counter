"""Excel export of analysis results in archive-ready format.

Mirrors the original ``For print.xlsx`` layout (one worksheet per vent
segment, manual measurements in the left block) and appends an automatic
measurement block on the right. Adds two summary worksheets:

  - **Methodology** — algorithm summary, DOI, author, generation date
  - **Summary**     — per-sheet accuracy table

Outlier rows (rotation mismatch) are highlighted in yellow.
"""
from __future__ import annotations

import os
from datetime import datetime

import io
import os
import tempfile

import openpyxl
import pandas as pd
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.trendline import Trendline
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FONT_HEADER = Font(name="Arial", size=11, bold=True)
FONT_BODY = Font(name="Arial", size=10)
FONT_META = Font(name="Arial", size=11, italic=True)
FONT_TITLE = Font(name="Arial", size=14, bold=True)

FILL_MANUAL = PatternFill("solid", start_color="DDEBF7")  # light blue
FILL_AUTO = PatternFill("solid", start_color="E2EFDA")    # light green
FILL_OUTLIER = PatternFill("solid", start_color="FFEB9C") # light yellow
FILL_HEADER = PatternFill("solid", start_color="305496")  # dark blue
FONT_HEADER_WHITE = Font(name="Arial", size=11, bold=True, color="FFFFFF")

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _autosize(ws, min_w: int = 10, max_w: int = 40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=min_w,
        )
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, length + 2))


def _write_methodology(wb, doi: str, summary_overall: dict, author: str, affiliation: str):
    ws = wb.create_sheet("Methodology", 0)
    ws["A1"] = "firic-rotation-counter — analysis archive"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:E1")

    rows = [
        ("Author", author),
        ("Affiliation", affiliation),
        ("Software", "firic-rotation-counter v0.1.0"),
        ("DOI", f"https://doi.org/{doi}"),
        ("Source code", "https://github.com/sooneeujuro/firic-rotation-counter"),
        ("Archive generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("Method", "Two-stage ROI (manual coarse + automatic flicker-based fine ROI), "
                  "HSV color-channel peak detection, autocorrelation-based period "
                  "estimation, smoke-occlusion-aware peak interpolation."),
        ("Frame rate", f"Read per video from the container (cv2.CAP_PROP_FPS); "
                       f"this dataset: {summary_overall.get('fps_str', 'n/a')}. "
                       f"Frame counts are converted to seconds with this rate "
                       f"(RPM is linear in fps)."),
        ("Validation", f"{summary_overall['matched']} "
                       f"({summary_overall['match_pct']:.1f}%) row-level rotation "
                       f"integer match against manual frame-by-frame counting; "
                       f"median |relative RPM error| = "
                       f"{summary_overall['median_rel_err']:.2f}%."),
        ("Limitations", "Fully smoke-occluded rotations are not extrapolated. "
                        "Sub-frame timing precision limits short-window measurements "
                        "to ~±1 frame."),
        ("", ""),
        ("Citation", f"Kim, H. (2026). firic-rotation-counter: automated rotation "
                     f"counting from ROV underwater flowmeter videos (Version 0.1.0) "
                     f"[Software]. Zenodo. https://doi.org/{doi}"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = FONT_HEADER
        ws.cell(row=i, column=2, value=v).font = FONT_BODY
        ws.cell(row=i, column=2).alignment = LEFT
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)

    ws.column_dimensions["A"].width = 18
    for col in "BCDE":
        ws.column_dimensions[col].width = 24
    for r in ws.iter_rows(min_row=3, max_row=3 + len(rows) - 1):
        for c in r:
            c.alignment = LEFT


def _write_summary(wb, summary_df: pd.DataFrame):
    ws = wb.create_sheet("Summary", 1)
    ws["A1"] = "Per-sheet validation summary"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:H1")

    cols = ["sheet", "indicator", "n_rows", "matched", "match_pct",
            "mean_rpm_err_pct", "mean_rpm_xl", "mean_rpm_auto"]
    headers = ["Sheet", "Indicator", "N rows", "Matched", "Match %",
               "Mean |Δ%|", "Mean RPM (manual)", "Mean RPM (auto)"]

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = FONT_HEADER_WHITE
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER

    for i, row in summary_df.iterrows():
        for j, col in enumerate(cols, start=1):
            v = row[col]
            c = ws.cell(row=4 + i, column=j, value=v)
            c.font = FONT_BODY
            c.alignment = CENTER
            c.border = BORDER
            if col in ("match_pct", "mean_rpm_err_pct", "mean_rpm_xl", "mean_rpm_auto") and isinstance(v, (int, float)):
                c.number_format = "0.00"
        if isinstance(row["match_pct"], (int, float)) and row["match_pct"] < 95:
            ws.cell(row=4 + i, column=5).fill = FILL_OUTLIER

    # overall row
    n_total = int(summary_df["n_rows"].sum())
    n_match = sum(int(s.split("/")[0]) for s in summary_df["matched"])
    n_valid = sum(int(s.split("/")[1]) for s in summary_df["matched"])
    overall_match = n_match / n_valid * 100
    overall_err = float(summary_df["mean_rpm_err_pct"].mean())
    last = 4 + len(summary_df)
    ws.cell(row=last, column=1, value="OVERALL").font = FONT_HEADER
    ws.cell(row=last, column=3, value=n_total).font = FONT_HEADER
    ws.cell(row=last, column=4, value=f"{n_match}/{n_valid}").font = FONT_HEADER
    ws.cell(row=last, column=5, value=overall_match).font = FONT_HEADER
    ws.cell(row=last, column=5).number_format = "0.00"
    ws.cell(row=last, column=6, value=overall_err).font = FONT_HEADER
    ws.cell(row=last, column=6).number_format = "0.00"
    for j in range(1, 9):
        ws.cell(row=last, column=j).fill = PatternFill("solid", start_color="D9E1F2")
        ws.cell(row=last, column=j).border = BORDER
        ws.cell(row=last, column=j).alignment = CENTER

    _autosize(ws, min_w=12, max_w=28)


def _write_sheet(wb, sheet_name: str, raw: pd.DataFrame, res: pd.DataFrame, indicator: str):
    """One vent worksheet: original metadata block + manual + auto columns."""
    ws = wb.create_sheet(sheet_name)

    # Row 1: metadata header
    headers_r1 = ["Dive #", "Vent ID", "Video #", "", "", "Indicator"]
    for j, h in enumerate(headers_r1, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = FONT_HEADER
        c.alignment = CENTER

    # Row 2: metadata values
    for j in range(6):
        v = raw.iloc[1, j] if j < raw.shape[1] else None
        if pd.notna(v):
            c = ws.cell(row=2, column=j + 1, value=v)
            c.font = FONT_META
            c.alignment = CENTER

    # Row 3: block titles
    ws.cell(row=3, column=1, value="── Manual measurement ──").font = FONT_HEADER
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    ws.cell(row=3, column=1).alignment = CENTER
    ws.cell(row=3, column=1).fill = FILL_MANUAL

    ws.cell(row=3, column=8, value="── Automatic measurement ──").font = FONT_HEADER
    ws.merge_cells(start_row=3, start_column=8, end_row=3, end_column=14)
    ws.cell(row=3, column=8).alignment = CENTER
    ws.cell(row=3, column=8).fill = FILL_AUTO

    # Row 4: column headers
    manual_h = ["Initial frame #", "Final frame #", "Time (s)", "Rotation", "RPM", "Timestamp"]
    auto_h = ["Rotation (auto)", "RPM (auto)", "ΔRPM", "|Δ%|", "Source", "n_R", "n_Y"]
    for j, h in enumerate(manual_h, start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = FONT_HEADER_WHITE
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    for j, h in enumerate(auto_h, start=8):
        c = ws.cell(row=4, column=j, value=h)
        c.font = FONT_HEADER_WHITE
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER

    # Data rows
    data = raw.iloc[3:, :6].reset_index(drop=True)
    for c in data.columns:
        data[c] = pd.to_numeric(data[c], errors="coerce")
    data = data.dropna(subset=[0, 1]).reset_index(drop=True)

    for i in range(len(data)):
        r = 5 + i
        for j in range(6):
            v = data.iloc[i, j]
            if pd.notna(v):
                cell = ws.cell(row=r, column=j + 1, value=float(v))
                cell.font = FONT_BODY
                cell.alignment = CENTER
                cell.border = BORDER
                if j in (0, 1, 3):
                    cell.number_format = "0"
                elif j == 2:
                    cell.number_format = "0.000000"
                else:
                    cell.number_format = "0.000000"

        if i < len(res):
            rr = res.iloc[i]
            auto_vals = [
                ("rot_auto", "0.0", 8),
                ("rpm_auto", "0.00", 9),
            ]
            for key, fmt, col in auto_vals:
                v = rr[key]
                if pd.notna(v):
                    cell = ws.cell(row=r, column=col, value=float(v))
                    cell.font = FONT_BODY
                    cell.alignment = CENTER
                    cell.border = BORDER
                    cell.number_format = fmt

            diff = float(rr["rpm_auto"] - rr["rpm_xl"]) if pd.notna(rr["rpm_auto"]) else None
            if diff is not None:
                c = ws.cell(row=r, column=10, value=diff)
                c.number_format = "+0.00;-0.00"
                c.font = FONT_BODY
                c.alignment = CENTER
                c.border = BORDER

            err_pct = float(rr["rpm_err_pct"]) if pd.notna(rr["rpm_err_pct"]) else None
            if err_pct is not None:
                c = ws.cell(row=r, column=11, value=abs(err_pct))
                c.number_format = "0.00"
                c.font = FONT_BODY
                c.alignment = CENTER
                c.border = BORDER

            for j, key in [(12, "src"), (13, "n_r"), (14, "n_y")]:
                v = rr.get(key)
                if pd.notna(v):
                    c = ws.cell(row=r, column=j, value=v)
                    c.font = FONT_BODY
                    c.alignment = CENTER
                    c.border = BORDER

            # outlier highlighting
            if pd.notna(rr["rot_xl"]) and pd.notna(rr["rot_auto"]):
                if round(rr["rot_auto"]) != rr["rot_xl"]:
                    for col in range(8, 15):
                        ws.cell(row=r, column=col).fill = FILL_OUTLIER

    # Footer note
    note_row = 5 + len(data) + 1
    n_match = int((res["rot_auto"].round() == res["rot_xl"]).sum())
    n_valid = int(res["rot_auto"].notna().sum())
    mean_err = float(res["rpm_err_pct"].abs().mean())
    c = ws.cell(
        row=note_row, column=1,
        value=f"Sheet summary: {n_match}/{n_valid} rotation match "
              f"({n_match / n_valid * 100:.1f}%), mean |Δ%| = {mean_err:.2f}%. "
              f"Yellow rows = rotation mismatch (typically smoke occlusion).",
    )
    c.font = FONT_META
    c.alignment = LEFT
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=14)

    # RPM time-series chart (ScatterChart so X axis is numeric and starts at 0)
    n_data = len(data)
    if n_data >= 2:
        valid_auto = res["rpm_auto"].dropna()
        if len(valid_auto) > 0:
            mean_auto = float(valid_auto.mean())
            # Helper column O (15): repeated mean (hidden) for chart line
            for i in range(n_data):
                c = ws.cell(row=5 + i, column=15, value=mean_auto)
                c.number_format = "0.00"
            ws.column_dimensions[get_column_letter(15)].hidden = True

            chart = ScatterChart()
            chart.title = sheet_name
            chart.style = 2
            chart.x_axis.title = "time (s)"
            chart.y_axis.title = "RPM"
            chart.x_axis.scaling.min = 0  # explicit X-axis start at 0
            chart.height = 9
            chart.width = 18
            chart.legend.position = "b"

            first, last = 5, 5 + n_data - 1
            x_ref = Reference(ws, min_col=6, min_row=first, max_row=last)
            auto_ref = Reference(ws, min_col=9, min_row=first, max_row=last)
            mean_ref = Reference(ws, min_col=15, min_row=first, max_row=last)

            # Auto line (with markers)
            s_auto = Series(values=auto_ref, xvalues=x_ref, title="RPM (auto)")
            s_auto.graphicalProperties = openpyxl.chart.shapes.GraphicalProperties(
                solidFill="1F77B4")
            s_auto.graphicalProperties.line.solidFill = "1F77B4"
            s_auto.graphicalProperties.line.width = 22000
            s_auto.smooth = False
            chart.series.append(s_auto)

            # Mean dashed line
            s_mean = Series(values=mean_ref, xvalues=x_ref,
                            title=f"mean = {mean_auto:.2f}")
            s_mean.graphicalProperties = openpyxl.chart.shapes.GraphicalProperties(
                solidFill="6BAED6")
            s_mean.graphicalProperties.line.solidFill = "6BAED6"
            s_mean.graphicalProperties.line.width = 14000
            s_mean.graphicalProperties.line.dashStyle = "dash"
            s_mean.smooth = False
            chart.series.append(s_mean)

            ws.add_chart(chart, f"P{first}")

    _autosize(ws, min_w=12, max_w=24)
    ws.freeze_panes = "A5"


def _embed_matplotlib_chart(ws, fig, anchor: str, scale: float = 1.0):
    """Render a matplotlib Figure to PNG and embed in the worksheet."""
    import matplotlib.pyplot as plt
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=110, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    img = XLImage(buf)
    if scale != 1.0:
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
    ws.add_image(img, anchor)


def export_archive_xlsx(
    src_xlsx_path: str,
    summary_df: pd.DataFrame,
    all_results: dict,
    out_path: str,
    doi: str,
    author: str,
    affiliation: str,
    rich: bool = False,
):
    """Write a full archive-ready xlsx.

    Parameters
    ----------
    src_xlsx_path : str
        Original manual-measurement spreadsheet (for metadata + manual rows).
    summary_df : pd.DataFrame
        Output of :func:`firic.pipeline.run_batch`.
    all_results : dict
        ``{sheet: (res_df, trace)}`` from run_batch.
    out_path : str
        Destination xlsx path.
    doi, author, affiliation : str
        For the Methodology sheet.
    """
    n_total = int(summary_df["n_rows"].sum())
    n_match = sum(int(s.split("/")[0]) for s in summary_df["matched"])
    n_valid = sum(int(s.split("/")[1]) for s in summary_df["matched"])
    combined = pd.concat(
        [r for r, _ in all_results.values()], ignore_index=True
    ).dropna(subset=["rpm_auto"])
    fps_values = sorted({round(float(t["fps"]), 3)
                         for _, t in all_results.values() if t.get("fps")})
    if not fps_values:
        fps_str = "n/a"
    elif len(fps_values) == 1:
        fps_str = f"{fps_values[0]:.3f} fps"
    else:
        fps_str = f"{fps_values[0]:.3f}–{fps_values[-1]:.3f} fps"
    summary_overall = {
        "n_rows": n_total,
        "matched": f"{n_match}/{n_valid}",
        "match_pct": n_match / n_valid * 100,
        "median_rel_err": float(combined["rpm_err_pct"].abs().median()),
        "fps_str": fps_str,
    }

    wb = openpyxl.Workbook()
    # remove default
    wb.remove(wb["Sheet"])

    _write_methodology(wb, doi, summary_overall, author, affiliation)
    _write_summary(wb, summary_df)

    for sheet, (res, trace) in all_results.items():
        raw = pd.read_excel(src_xlsx_path, sheet_name=sheet, header=None)
        _write_sheet(wb, sheet, raw, res, trace["indicator"])

    # Embed a small per-vent summary chart on the Summary sheet (only meaningful for N≥2)
    if len(all_results) >= 2:
        from .viz import plot_vent_summary
        ws_sum = wb["Summary"]
        anchor_row = 4 + len(summary_df) + 3
        fig = plot_vent_summary(all_results)
        if fig is not None:
            _embed_matplotlib_chart(ws_sum, fig, f"A{anchor_row}", scale=0.85)

        if rich:
            from .viz import plot_vent_boxplot, plot_agreement
            anchor_row2 = anchor_row + 28
            fig2 = plot_vent_boxplot(all_results)
            if fig2 is not None:
                _embed_matplotlib_chart(ws_sum, fig2, f"A{anchor_row2}", scale=0.85)
            anchor_row3 = anchor_row2 + 28
            combined = pd.concat(
                [r for r, _ in all_results.values()], ignore_index=True
            ).dropna(subset=["rpm_auto"])
            if "sheet" not in combined.columns:
                # rebuild sheet column from order
                combined = pd.concat(
                    [r.assign(sheet=s) for s, (r, _) in all_results.items()],
                    ignore_index=True,
                ).dropna(subset=["rpm_auto"])
            import matplotlib.pyplot as plt
            fig3, axes = plt.subplots(1, 2, figsize=(12, 5))
            df = combined.copy()
            df["rpm_diff"] = df["rpm_auto"] - df["rpm_xl"]
            axes[0].scatter(df["rpm_xl"], df["rpm_auto"], alpha=0.5, s=20)
            lim = [df[["rpm_xl", "rpm_auto"]].min().min() - 5,
                   df[["rpm_xl", "rpm_auto"]].max().max() + 5]
            axes[0].plot(lim, lim, "k--", alpha=0.5, label="y = x")
            axes[0].set_xlabel("Manual RPM")
            axes[0].set_ylabel("Auto RPM")
            axes[0].set_title(f"Manual vs Auto (n = {len(df)})")
            axes[0].legend()
            axes[0].grid(alpha=0.3)
            axes[0].set_aspect("equal")
            mean_rpm = (df["rpm_xl"] + df["rpm_auto"]) / 2
            axes[1].scatter(mean_rpm, df["rpm_diff"], alpha=0.5, s=20)
            axes[1].axhline(df["rpm_diff"].mean(), color="r",
                            label=f"bias = {df['rpm_diff'].mean():.2f}")
            axes[1].axhline(df["rpm_diff"].mean() + 1.96 * df["rpm_diff"].std(),
                            color="r", ls="--", alpha=0.5)
            axes[1].axhline(df["rpm_diff"].mean() - 1.96 * df["rpm_diff"].std(),
                            color="r", ls="--", alpha=0.5)
            axes[1].set_xlabel("Mean RPM")
            axes[1].set_ylabel("Auto − Manual")
            axes[1].set_title("Bland-Altman agreement")
            axes[1].legend()
            axes[1].grid(alpha=0.3)
            plt.tight_layout()
            _embed_matplotlib_chart(ws_sum, fig3, f"A{anchor_row3}", scale=0.85)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path
